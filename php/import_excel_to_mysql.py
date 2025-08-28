import openpyxl
import pymysql
from pymysql.cursors import DictCursor

# Database configuration
db_config = {
    'host': 'mysql2103.db.sakura.ne.jp',  # Apna MySQL host
    'user': 'it-future_itf',  # Apna username
    'password': 'future1800',  # Apna password
    'database': 'it-future_itf',  # Apna database name
    'port': 3306,  # Default, verify with Sakura
    'cursorclass': DictCursor,
    'connect_timeout': 10  # Add timeout
}

# Excel file path
excel_file = '(ITF)顧客入力リスト(更新版2024.02～）.xlsx'

# MySQL table columns
columns = [
    '採用日時', '施設名（勤務先）', '雇用者情報（アルファベット）', '管理番号', '担当者（企業）',
    '基本契約書', '委託契約書', '紹介元', '受入機関（郵便番号）', '受入機関（住所）',
    '請求書送付先', '受入機関（電話番号）', '担当責任者', '区分', '受入機関名（所属機関）',
    '雇用者情報（カタカナ）', '雇用者情報（性別）', '雇用者情報（国籍）', '雇用者情報（生年月日）',
    '年齢', '雇用者在留番号', '雇用者在留期限', '更新回数', 'X', '入社日', '在留カード最初発行日',
    '支援退職日', '状態', '管理費', '紹介料', '住居タイプ', '不動産会社', '不動産連絡先', '支援者住所',
    '連絡先①', 'AJ', 'AK', 'AL', 'AM', '支援者の家賃', '共益費', 'AP', '満了時期', '備考欄',
    '正担当者', 'JLPT', 'エリア', '受け入れ期間', '紹介手数料', '四半期', 'AY', 'AZ', 'BA',
    'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH'
]

try:
    # Connect to MySQL
    print("Attempting to connect to MySQL...")
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    print("Connection successful!")

    # Load Excel workbook
    print(f"Loading Excel file: {excel_file}")
    workbook = openpyxl.load_workbook(excel_file)

    # Loop through sheets
    for sheet_name in workbook.sheetnames:
        if sheet_name not in ["定義", "支援状況"]:  # Skip non-data sheets
            sheet = workbook[sheet_name]
            print(f"Processing sheet: {sheet_name}")

            # Skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Clean and prepare data
                row_data = [cell if cell is not None else '' for cell in row[:len(columns)]]
                row_data += [''] * (len(columns) - len(row_data))  # Pad with empty strings

                # Create INSERT query
                query = f"INSERT INTO talents ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(columns))})"
                try:
                    cursor.execute(query, row_data)
                    connection.commit()
                    print(f"Inserted row from {sheet_name}: {row_data[2]}")
                except Exception as e:
                    print(f"Error inserting row from {sheet_name}: {e}")
                    connection.rollback()

    print("Import completed!")
except pymysql.Error as e:
    print(f"Database error: {e}")
except Exception as e:
    print(f"General error: {e}")
finally:
    if 'connection' in locals():
        cursor.close()
        connection.close()
        print("Connection closed.")